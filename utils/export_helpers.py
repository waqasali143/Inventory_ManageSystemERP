import os
import tempfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def export_to_excel(headers, rows, filename_prefix):
    """
    headers: list of column names, e.g. ["Sale No", "Customer", "Date", "Net Total"]
    rows: list of tuples matching those headers
    filename_prefix: used to name the file, e.g. "Sales_History"

    Creates the .xlsx file in the system temp folder and opens it.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # ---- Header row ----
    ws.append(headers)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # ---- Data rows ----
    for row in rows:
        ws.append(row)

    # ---- Auto-size columns (roughly) ----
    for col_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
        col_letter = col_cells[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 4

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(tempfile.gettempdir(), f"{filename_prefix}_{timestamp}.xlsx")
    wb.save(file_path)

    os.startfile(file_path)

    return file_path
# ==========================================================================

def export_multi_sheet_excel(sheets, filename_prefix):
    """
    sheets: list of (sheet_name, headers, rows) tuples.
    Creates one Excel file with multiple sheets - one per tuple.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove the default blank sheet

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet_name, headers, rows in sheets:
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name limit

        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in rows:
            ws.append(row)

        for col_cells in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
            col_letter = col_cells[0].column_letter
            ws.column_dimensions[col_letter].width = max_length + 4

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(tempfile.gettempdir(), f"{filename_prefix}_{timestamp}.xlsx")
    wb.save(file_path)

    os.startfile(file_path)

    return file_path