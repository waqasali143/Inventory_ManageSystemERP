
from tkinter import messagebox
from repositories import product_repository as repo
from utils import event_bus
from openpyxl import Workbook, load_workbook

# =====================================================================
# GET PRODUCT STATUS
# =====================================================================
def get_product_status(quantity):
    """
    Quantity = 0      -> Out of Stock
    Quantity 1-5      -> Low Stock
    Quantity > 5      -> Active
    """
    if quantity <= 0:
        return "Out of Stock"
    elif quantity <= 5:
        return "Low Stock"
    else:
        return "Active"

# =====================================================================
# VALIDATE PRODUCT DATA
# =====================================================================
def validate_product_data(name, cost_price, sale_price, quantity):
    """
    Returns:
        (True, "")                -> Validation Passed
        (False, "Error Message")  -> Validation Failed
    """
    product_name = name.get().strip()

    if product_name == "":
        return False, "Product Name is required."

    if len(product_name) < 2:
        return False, "Product Name must contain at least 2 characters."

    if len(product_name) > 100:
        return False, "Product Name cannot exceed 100 characters."

    try:
        product_cost = float(cost_price.get())
    except ValueError:
        return False, "Invalid Cost Price."

    if product_cost <= 0:
        return False, "Cost Price must be greater than zero."

    try:
        product_sale = float(sale_price.get())
    except ValueError:
        return False, "Invalid Sale Price."

    if product_sale <= 0:
        return False, "Sale Price must be greater than zero."

    if product_sale < product_cost:
        return False, "Sale Price cannot be less than Cost Price."

    try:
        product_qty = int(quantity.get())
    except ValueError:
        return False, "Quantity must be an integer."

    if product_qty < 0:
        return False, "Quantity cannot be negative."

    return True, ""

# =====================================================================
# LOAD / SEARCH PRODUCTS
# =====================================================================
def load_products(search_term=None):
    return repo.fetch_products(search_term)

# =====================================================================
# SAVE PRODUCT
# =====================================================================
def save_product(name, cost_price, sale_price, quantity, barcode=None):

    is_valid, message = validate_product_data(name, cost_price, sale_price, quantity)
    if not is_valid:
        messagebox.showerror("Validation Error", message)
        return False

    product_name = name.get().strip()
    barcode_value = barcode.get().strip() if barcode else ""

    if repo.fetch_product_by_name(product_name):
        messagebox.showerror("Error", "Product already exists!")
        return False

    if barcode_value and repo.fetch_product_by_barcode_unique_check(barcode_value):
        messagebox.showerror("Error", "This barcode is already assigned to another product.")
        return False

    repo.insert_product(
        product_name,
        float(cost_price.get()),
        float(sale_price.get()),
        int(quantity.get()),
        get_product_status(int(quantity.get())),
        barcode_value
    )

    event_bus.publish()

    messagebox.showinfo("Success", "Product Added Successfully")
    return True

# =====================================================================
# UPDATE PRODUCT
# =====================================================================
def update_product(selected_id, name, cost_price, sale_price, quantity, barcode=None):

    if not selected_id:
        messagebox.showerror("Error", "Please select a product first.")
        return False

    is_valid, message = validate_product_data(name, cost_price, sale_price, quantity)
    if not is_valid:
        messagebox.showerror("Validation Error", message)
        return False

    product_name = name.get().strip()
    barcode_value = barcode.get().strip() if barcode else ""

    if repo.fetch_product_by_name(product_name, exclude_id=selected_id):
        messagebox.showerror("Duplicate Product", "Product name already exists.")
        return False

    if barcode_value and repo.fetch_product_by_barcode_unique_check(barcode_value, exclude_id=selected_id):
        messagebox.showerror("Error", "This barcode is already assigned to another product.")
        return False

    product_quantity = int(quantity.get())

    repo.update_product(
        selected_id,
        product_name,
        float(cost_price.get()),
        float(sale_price.get()),
        product_quantity,
        get_product_status(product_quantity),
        barcode_value
    )

    event_bus.publish()

    messagebox.showinfo("Success", "Product Updated Successfully")
    return True

# =====================================================================
# DELETE PRODUCT
# (falls back to deactivating the product if it's referenced by a
#  past purchase/sale - hard deleting it would break that history)
# =====================================================================
def delete_product(selected_id):

    if not selected_id:
        messagebox.showerror("Error", "Please select a product first.")
        return False

    answer = messagebox.askyesno(
        "Delete Product",
        "Are you sure you want to delete this product?"
    )
    if not answer:
        return False

    success, reason = repo.delete_product(selected_id)

    if success:
        event_bus.publish()
        messagebox.showinfo("Success", "Product Deleted Successfully")
        return True

    if reason == "in_use":
        deactivate = messagebox.askyesno(
            "Cannot Delete",
            "This product is used in existing purchase/sale records, "
            "so it can't be deleted.\n\n"
            "Do you want to mark it as Inactive instead? "
            "(It will be hidden from new purchases/sales, but its "
            "history stays intact.)"
        )
        if deactivate:
            repo.set_product_status(selected_id, "Inactive")
            event_bus.publish()
            messagebox.showinfo("Success", "Product marked as Inactive.")
            return True

    return False

# =====================================================================
# BULK IMPORT (Excel)
# =====================================================================
def create_import_template(file_path):
    """Generates a blank Excel template with the correct headers."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    ws.append(["Product Name", "Cost Price", "Sale Price", "Quantity"])

    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    ws.append(["Example Product", 500, 750, 20])

    wb.save(file_path)


def import_products_from_excel(file_path):
    """
    Reads the uploaded Excel file, validates each row, and inserts
    valid products in bulk.

    Returns (success_count, skipped_names, error_rows)
        error_rows: list of (row_number, reason) for bad data
    """
    wb = load_workbook(file_path)
    ws = wb.active

    valid_products = []
    error_rows = []

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        if row is None or all(cell is None for cell in row):
            continue  # skip fully blank rows

        name, cost_price, sale_price, quantity = (list(row) + [None, None, None, None])[:4]

        if not name or str(name).strip() == "":
            error_rows.append((row_number, "Product Name is missing."))
            continue

        try:
            cost_price = float(cost_price)
            sale_price = float(sale_price)
            quantity = int(quantity)
        except (TypeError, ValueError):
            error_rows.append((row_number, "Cost Price / Sale Price / Quantity must be numbers."))
            continue

        if cost_price <= 0 or sale_price <= 0:
            error_rows.append((row_number, "Cost Price and Sale Price must be greater than zero."))
            continue

        if quantity < 0:
            error_rows.append((row_number, "Quantity cannot be negative."))
            continue

        status = get_product_status(quantity)
        valid_products.append((str(name).strip(), cost_price, sale_price, quantity, status))

    inserted, skipped = repo.bulk_insert_products(valid_products)

    if inserted > 0:
        event_bus.publish()

    return inserted, skipped, error_rows
# ===============================================================

def get_product_by_barcode(barcode):
    return repo.fetch_product_by_barcode(barcode)